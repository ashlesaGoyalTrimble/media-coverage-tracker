"""Media-related services for processing articles and images."""
import requests
from typing import List
import io
import asyncio
import httpx
import traceback
import pandas as pd
import time
from bs4 import BeautifulSoup
import openpyxl
from io import BytesIO
import os
from urllib.parse import urljoin
import pytesseract
from PIL import Image
import tempfile
import base64
import re
import uuid
from selenium import webdriver
from fastapi.responses import StreamingResponse
from openpyxl.utils.dataframe import dataframe_to_rows
from fastapi import HTTPException, UploadFile, File

from app.schemas.media import MessageRequest, AssistantCreateRequest, Tool
from app.services.auth_service import get_trimble_auth_headers


BASE_URL = "https://api.assistant.trimble.cloud/ui/trimbledeveloperprogram/assistants/v1"
API_ENDPOINT = "http://localhost:8000/agents/all/messages"

ASSISTANT_MAP = {
    "Content-type": "trimble-media-content-type",
    "Corporate": "trimble-media-corporate",
    "Media-types": "trimble-media-media-types",
    "Field-systems": "trimble-media-field-systems",
    "aeco": "trimble-media-coverage-aeco",
}

categories = [
    "Publication", "Article Title & Link", "Date", "Qtr", "Country", "Global Region Reached",
    "Corporate", "AECO", "B2W", "MEP", "SketchUp Visualization", "SketchUp Collaboration",
    "Structures", "Viewpoint", "Industry Cloud/TC1", "Civil Design & Engineering",
    "Civil Construction (CIS)", "O&PS", "FIELD SYSTEMS", "Civil", "Geospatial / BCFS",
    "Applanix", "OEM GNSS", "TAP / Auto IoT", "Paving / Milling", "Marine", "Drilling / Piling",
    "Earthmoving / Machine Control", "Surveying (human / drone / machine)", 
    "Bidding / Estimating / Takeoff", "Jobsite connectivity / F2O", "Safety",
    "Asset capture and inspection", "Monitoring", "Reality capture",
    "BIM / Model-based workflows", "Mixed reality", "Crash & Crime",
    "Field Systems Themes", "TRANSPORTATION & LOG.", "Forestry", "Mobility",
    "Transporeon", "MAPS", "Rail", "Thought Leadership / Byline", "Journalist Feature",
    "Customer Focus", "Award", "Podcast", "News release pickup", "Mention", "GREAT ONE",
    "Trimble in video", "Trimble quote", "Trimble image", "Trimble title mention",
    "T1: Business / Finance", "T1: Dailies", "T1: TV/Radio", "T1: Technology", "T1: Industry",
    "T2: Dailies, Business, Regional", "T2: Trade", "T2: Technology", "T2: Industry (adjacent)",
    "AI/ML", "Infrastructure", "Trimble revenue / business growth", 
    "Digital 2 Physical / Ph2Dig", "Connected Ecosystems", "Sustainability",
    "Trust & Security", "Workforce Optimization", "Innovation"]


# Function to read hyperlinks from Excel
async def read_hyperlinks(file: UploadFile, sheet_name: str) -> pd.DataFrame:
    """Reads URLs from an uploaded Excel file."""
    try:
        contents = await file.read()
        return pd.read_excel(BytesIO(contents), sheet_name=sheet_name)
    except Exception as e:
        print(f"Exception in read_hyperlinks: {e}")
        return pd.DataFrame()


# Function to extract text from web articles
def web_scrape_text(url: str) -> str:
    """Extracts article text from the given URL."""
    try:
        # Add short timeout to prevent hanging
        response = requests.get(
            url, 
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}, 
            timeout=15  # 15 second timeout
        )
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            text = soup.get_text().strip()
            return text if text else "No content found"
        else:
            print(f"HTTP {response.status_code} for {url}")
            return f"Failed to scrape URL: {url} (HTTP {response.status_code})"
    except requests.exceptions.Timeout:
        print(f"Timeout scraping {url}")
        return f"Error: Timeout scraping {url}"
    except requests.exceptions.ConnectionError as e:
        print(f"Connection error scraping {url}: {e}")
        return f"Error: Connection failed for {url}"
    except Exception as e:
        print(f"Exception in web_scrape_text for {url}: {e}")
        return f"Error: Failed to scrape {url}"


# Function to call the assistant API with the scraped text
async def call_assistant(url: str, headers: dict, payload: dict) -> str:
    """Sends a message to the assistant API with timeout and retry logic."""
    try:
        for attempt in range(2):  # Reduced to 2 attempts for faster processing
            try:
                # Shorter timeout to prevent AWS gateway timeouts
                async with httpx.AsyncClient(timeout=20.0) as client:
                    response = await client.post(url, headers=headers, json=payload)
                    
                    if response.status_code == 200:
                        try:
                            json_data = response.json()
                            return json_data.get("message", str(json_data))
                        except Exception as json_e:
                            print(f"JSON parse error on attempt {attempt + 1}: {json_e}")
                            return response.text if response.text else "Empty response"
                    elif response.status_code == 504:
                        print(f"504 Gateway timeout on attempt {attempt + 1} for {url}")
                        if attempt < 1:
                            await asyncio.sleep(1)
                            continue
                        else:
                            return "Error: Gateway timeout (504)"
                    else:
                        print(f"HTTP {response.status_code} on attempt {attempt + 1} for {url}")
                        return f"Error: HTTP {response.status_code}"
                        
            except (httpx.TimeoutException, httpx.ReadTimeout) as timeout_e:
                print(f"Timeout on attempt {attempt + 1} for {url}: {timeout_e}")
                if attempt < 1:
                    await asyncio.sleep(1)
                    continue
                else:
                    return "Error: Request timeout"
                    
            except (httpx.NetworkError, httpx.ConnectError) as net_e:
                print(f"Network error on attempt {attempt + 1} for {url}: {net_e}")
                if attempt < 1:
                    await asyncio.sleep(1)
                    continue
                else:
                    return f"Error: Network error - {net_e}"
                    
            except Exception as e:
                error_msg = str(e)
                print(f"Error on attempt {attempt + 1} for {url}: {error_msg}")
                if "504" in error_msg or "Gateway" in error_msg:
                    if attempt < 1:
                        await asyncio.sleep(1)
                        continue
                    else:
                        return "Error: Gateway timeout"
                else:
                    return f"Error: {error_msg}"
                    
        return "Error: All attempts failed"
    except Exception as e:
        print(f"Fatal error in call_assistant for {url}: {e}")
        return f"Error: {e}"


# Function to upload an image and get its blob URL
async def upload_image(assistant_id: str, session_id: str, file: UploadFile) -> dict:
    """Uploads an image to the assistant and returns the response."""
    try:
        url = f"{BASE_URL}/agents/{assistant_id}/sessions/{session_id}/images"
        files = {"file": (file.filename, await file.read(), file.content_type)}
        
        # Get fresh authentication headers
        headers = await get_trimble_auth_headers()
        # Remove Content-Type for multipart/form-data uploads
        auth_headers = {k: v for k, v in headers.items() if k != "Content-Type"}
        
        response = requests.post(url, headers=auth_headers, files=files)
        
        if response.status_code == 200:
            return response.json()
        else:
            return {"error": f"Upload failed with status code {response.status_code}"}
    except Exception as e:
        print(f"Exception in upload_image: {e}")
        return {"error": "catched error"}


# Function to check if a URL is an image
def is_image_url(url: str) -> bool:
    """Checks whether the given URL is likely to be an image."""
    try:
        return "qg" in url.lower() or "digital" in url.lower()
    except Exception as e:
        print(f"Exception in is_image_url: {e}")
        return False

# Writes the output to the Excel file in memory
def write_output_to_excel_memory(file_contents: bytes, df: pd.DataFrame) -> bytes:
    """Writes the output to the Excel file and returns the updated file as bytes."""
    try:
        wb = openpyxl.load_workbook(BytesIO(file_contents))
        sheet = wb["MediaScorecard"]  # Updated sheet name

        # Fixed known column range: CORP. (column 13) to CORPORATE THEMES (column 75)
        start_col_idx = 13
        end_col_idx = 75

        # Clear existing "X" marks from category columns (from row 5 to end)
        for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):
            for col_idx in range(start_col_idx, end_col_idx + 1):
                row[col_idx - 1].value = None

        # Write updated "X" values from the dataframe (assumes df starts with same headers)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=5):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # Save to BytesIO and return the bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    except Exception as e:
        print(f"Exception in write_output_to_excel_memory: {e}")
        return file_contents  # Return original file if processing fails

# Function to process an image link and return its blob URL
async def process_image_link(url: str) -> str:
    """Uploads image to the assistant and returns its blob URL."""
    try:
        response = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})
        if response.status_code == 200:
            file_content = BytesIO(response.content)
            file = UploadFile(filename=url.split("/")[-1], file=file_content)
            upload_response = await upload_image("trimble-media-image-2-text", str(uuid.uuid4()), file)
            return upload_response.get("blob_url", "No blob URL found")
        else:
            return f"Failed to retrieve image: {response.status_code}"
    except Exception as e:
        print(f"Exception in process_image_link: {e}")
        return "Error: catched error"


# Sends a message to a specific assistant and returns the response
async def send_message(assistant_id: str, request: MessageRequest):
    try:
        url = f"{BASE_URL}/agents/{assistant_id}/messages"
        
        # Get fresh authentication headers
        headers = await get_trimble_auth_headers()
        payload = request.dict()
        
        # Add timeout to prevent hanging
        response = requests.post(url, headers=headers, json=payload, timeout=25)
        
        if response.status_code == 200:
            try:
                return response.json()
            except Exception as e:
                print(f"JSON parsing error in send_message: {e}")
                return {"message": "Error: JSON parsing failed"}
        elif response.status_code == 504:
            print(f"504 Gateway timeout for assistant {assistant_id}")
            return {"message": "Error: Gateway timeout (504)"}
        else:
            print(f"HTTP {response.status_code} for assistant {assistant_id}: {response.text}")
            return {"message": f"Error: HTTP {response.status_code}"}
            
    except requests.exceptions.Timeout:
        print(f"Request timeout for assistant {assistant_id}")
        return {"message": "Error: Request timeout"}
    except requests.exceptions.ConnectionError as e:
        print(f"Connection error for assistant {assistant_id}: {e}")
        return {"message": "Error: Connection failed"}
    except Exception as e:
        error_msg = str(e)
        print(f"Exception in send_message for {assistant_id}: {error_msg}")
        if "504" in error_msg or "Gateway" in error_msg:
            return {"message": "Error: Gateway timeout"}
        else:
            return {"message": f"Error: {error_msg}"}


# Sends a message to all assistants and consolidates the responses
async def send_to_all_assistants(request: MessageRequest):
    try:
        # Get fresh authentication headers
        headers = await get_trimble_auth_headers()
        payload = request.dict()
        tasks = [
            call_assistant(f"{BASE_URL}/agents/{ASSISTANT_MAP[aid]}/messages", headers, payload)
            for aid in ASSISTANT_MAP
        ]

        results = await asyncio.gather(*tasks, return_exceptions=True)

        # Handle exceptions in results
        processed_results = []
        for aid, result in zip(ASSISTANT_MAP.keys(), results):
            if isinstance(result, Exception):
                print(f"Exception from assistant {aid}: {result}")
                processed_results.append(f"{aid}: Error - {str(result)}")
            else:
                processed_results.append(f"{aid}: {result}")

        consolidated = "\n".join(processed_results)
        return {"consolidated_response": consolidated}
    except Exception as e:
        print(f"Exception in send_to_all_assistants: {e}")
        return {"consolidated_response": "Error: catched error"}


# Main function to extract text, send to API, and store X for matching categories

async def process_hyperlinks(file: UploadFile, sheet_name: str) -> bytes:
    """Process hyperlinks from uploaded Excel file and return processed file as bytes."""
    try:
        # Read the original file contents
        file_contents = await file.read()
        
        # Reset file pointer for reading the dataframe
        await file.seek(0)
        df = await read_hyperlinks(file, sheet_name)

        async def process_link(row):
            """Handles individual link processing asynchronously."""
            try:
                link = row.get("Unnamed: 1")
                if pd.notnull(link):
                    print(f"Processing link: {link}")

                    response_text = "Error: No response"
                    try:
                        if is_image_url(link):
                            # Process image URL with specific error handling
                            print(f"Processing as image: {link}")
                            try:
                                # Add timeout for image processing
                                blob_url = await asyncio.wait_for(
                                    process_image_link(link), 
                                    timeout=30
                                )
                                print(f"Got blob URL for {link}: {blob_url[:50]}..." if blob_url else "No blob URL")
                                
                                message_request = MessageRequest(message=blob_url, stream=False)
                                assistant_response = await asyncio.wait_for(
                                    send_message("trimble-media-image-2-text", message_request),
                                    timeout=30
                                )
                                response_text = assistant_response.get("message", "No text found")
                                print(f"Assistant response for image {link}: Success")
                            except asyncio.TimeoutError:
                                print(f"Timeout in image processing for {link}")
                                response_text = "Error: Image processing timeout"
                            except Exception as img_e:
                                error_msg = str(img_e)
                                print(f"Exception in image processing for {link}: {error_msg}")
                                if "504" in error_msg or "timeout" in error_msg.lower():
                                    response_text = "Error: Image processing timeout"
                                else:
                                    response_text = "Error: Image processing failed"
                        else:
                            # Process article URL with specific error handling
                            print(f"Processing as article: {link}")
                            try:
                                # Add timeout for web scraping
                                scraped_text = await asyncio.wait_for(
                                    asyncio.get_event_loop().run_in_executor(
                                        None, 
                                        web_scrape_text, 
                                        link
                                    ),
                                    timeout=20
                                )
                                print(f"Scraped text length for {link}: {len(scraped_text)} characters")
                                
                                message_request = MessageRequest(message=scraped_text, stream=False)
                                assistant_response = await asyncio.wait_for(
                                    send_to_all_assistants(message_request),
                                    timeout=45  # Longer timeout for multiple assistants
                                )
                                response_text = assistant_response.get("consolidated_response", "No response")
                                print(f"Assistant response for article {link}: Success")
                            except asyncio.TimeoutError:
                                print(f"Timeout in article processing for {link}")
                                response_text = "Error: Article processing timeout"
                            except Exception as article_e:
                                error_msg = str(article_e)
                                print(f"Exception in article processing for {link}: {error_msg}")
                                if "504" in error_msg or "timeout" in error_msg.lower():
                                    response_text = "Error: Article processing timeout"
                                else:
                                    response_text = "Error: Article processing failed"
                                
                    except Exception as e:
                        print(f"Exception in processing link {link}: {e}")
                        response_text = "Error: Link processing failed"
                    
                    # Always continue to category processing, even if assistant calls failed
                    print(f"Continuing to category processing for {link} with response: {response_text[:100]}...")

                    # Always create a result row, regardless of what happened above
                    try:
                        row_dict = {cat: "" for cat in categories}
                        row_dict["Article Title & Link"] = link

                        # Only process categories if we have meaningful response text
                        if response_text and not response_text.startswith("Error:") and len(response_text.strip()) > 10:
                            category_count = 0
                            for category in categories:
                                try:
                                    if re.search(rf'\b{re.escape(category)}\b', response_text, flags=re.IGNORECASE):
                                        row_dict[category] = "X"
                                        category_count += 1
                                except Exception as regex_e:
                                    print(f"Exception in regex for category {category}: {regex_e}")
                                    continue
                            print(f"Found {category_count} matching categories for {link}")
                        else:
                            print(f"No valid response text for category matching: {response_text}")

                        print(f"Successfully processed {link} - moving to next link")
                        return row_dict
                        
                    except Exception as e:
                        print(f"Exception in processing categories for {link}: {e}")
                        # Still return a basic row to keep processing going
                        row_dict = {cat: "" for cat in categories}
                        row_dict["Article Title & Link"] = link
                        print(f"Returning basic row for {link} due to category processing error")
                        return row_dict

                return None
            except Exception as e:
                print(f"Exception in process_link for {link if 'link' in locals() else 'unknown link'}: {e}")
                # Even on complete failure, try to return something to continue processing
                try:
                    link = row.get("Unnamed: 1", "unknown")
                    row_dict = {cat: "" for cat in categories}
                    row_dict["Article Title & Link"] = str(link)
                    print(f"Returning error row for {link} to continue processing")
                    return row_dict
                except Exception:
                    print(f"Could not create error row, skipping this link entirely")
                    return None

        # Process all links with maximum resilience and controlled concurrency
        total_links = len(df)
        print(f"Starting to process {total_links} links with full error resilience...")
        
        # Limit concurrency to prevent AWS gateway timeouts
        semaphore = asyncio.Semaphore(5)  # Max 5 concurrent requests
        
        async def process_link_with_semaphore(row):
            async with semaphore:
                return await process_link(row)
        
        try:
            tasks = [process_link_with_semaphore(row) for _, row in df.iterrows()]
            # Use return_exceptions=True to ensure NO task failure stops others
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            # Process results with detailed error reporting
            valid_results = []
            error_count = 0
            success_count = 0
            
            for i, result in enumerate(results):
                if isinstance(result, Exception):
                    error_count += 1
                    print(f"Task {i+1}/{total_links} failed with exception: {result}")
                    # Continue to next task without stopping
                    continue
                elif result is not None:
                    success_count += 1
                    valid_results.append(result)
                    print(f"Task {i+1}/{total_links} completed successfully")
                else:
                    print(f"Task {i+1}/{total_links} returned None (skipped)")
            
            print(f"\nProcessing Summary:")
            print(f"  Total links: {total_links}")
            print(f"  Successful: {success_count}")
            print(f"  Errors: {error_count}")
            print(f"  Skipped: {total_links - success_count - error_count}")
            
            result_df = pd.DataFrame(valid_results) if valid_results else pd.DataFrame()
            print(f"  Final DataFrame has {len(result_df)} rows")
            
        except Exception as e:
            print(f"Exception in gathering results: {e}")
            print("Creating empty DataFrame due to gathering failure")
            result_df = pd.DataFrame()

        # Write output to Excel file in memory and return the bytes
        processed_file_bytes = write_output_to_excel_memory(file_contents, result_df)
        
        return processed_file_bytes
    except Exception as e:
        print(f"Exception in process_hyperlinks: {e}")
        # Return original file if everything fails
        try:
            file_contents = await file.read()
            return file_contents
        except Exception as file_e:
            print(f"Exception reading original file: {file_e}")
            return b""